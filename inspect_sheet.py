import openpyxl, glob, os

# xlsxファイルを自動検索
files = glob.glob(r"C:\Users\Loser\Desktop\-\tamalabo\*.xlsx")
print("見つかったxlsxファイル:", files)
if not files:
    print("xlsxファイルが見つかりません")
    exit(1)

wb = openpyxl.load_workbook(files[0], data_only=True)
ws = wb.active
print("シート名:", ws.title)
print("最大行:", ws.max_row)
print("最大列:", ws.max_column)
print()

# AB=28, BC=55
print("=== AB列(28) ～ BC列(55) の1行目(ヘッダー) ===")
for col in range(28, 56):
    cell = ws.cell(row=1, column=col)
    col_letter = openpyxl.utils.get_column_letter(col)
    print("  %s(%d): %s" % (col_letter, col, cell.value))

print()
print("=== AB～BC列 の全データ行 ===")
for row in range(2, ws.max_row + 1):
    vals = []
    for col in range(28, 56):
        vals.append(str(ws.cell(row=row, column=col).value))
    print("行%d: %s" % (row, ",".join(vals)))
